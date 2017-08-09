import json
import pprint
import re
import datetime
import openpyxl
import copy

from openpyxl.styles import Font
from openpyxl.styles.colors import RED
font = Font(color=RED)

IS_SCRAPING_ESPN = True
# ================================================================================================================

n_previous_games = 4
n_games_per_sheet = 8

# ================================================================================================================

ratings_data_column_start = ord("l") - 96 - 1
col_offset_home_away = 2
col_offset_games = 4

header_row = 9
result_row = 10
runs_row = 11

# ================================================================================================================

profile_data_row_start = 5
profile_data_column_start = ord("d") - 96
profile_neat_layout_column_jump = (ord("q") - 96) - profile_data_column_start  # second set of games starts at col Q

profile_row_offset_home_away = 1
profile_row_offset_game = 3

# ================================================================================================================

suitability_stats_col_offset = 1
suitability_home_away_col_additional_offset = 1

matchup_stats_col_start_offset = 1
matchup_stats_col_offset = 1
matchup_stats_row = 39


batting_total_row = 22
batting_home_row = 23
batting_away_row = 24

team_home_row = 27
team_away_row = 28

pitcher_home_row = 31
pitcher_away_row = 32

# ================================================================================================================

win_loss_key = "W/L"
team_score_key = "Score"
pitcher_name_key = 'name'
pitcher_win_loss_key = "RESULT" if IS_SCRAPING_ESPN else "DECISION"

# =============================================sub function =========================================================

def WriteHeader(team_data, worksheet, col_offset=0, is_home=None):

    if is_home is None:
        home_away = "TEAM: "
    elif is_home:
        home_away = "HOME: "
    else:
        home_away = "AWAY: "

    worksheet.cell(column=col_offset, row=header_row, value="{} {}".format(home_away, team_data['team_id']))

def WriteResult(team_data, worksheet, col_offset=0):
    # wins and losses
    for i in range(n_previous_games):
        reverse_excel_idx = n_previous_games - i - 1
        # REVERSE DATA INDEX, scraped table data is ascending chronological & we want most recent
        W_or_L = team_data[ win_loss_key ][-1-i]
        worksheet.cell(column=col_offset + reverse_excel_idx, row=result_row, value= W_or_L)

def WriteRuns(team_data, worksheet, col_offset=0):
    # numbers
    for i in range(n_previous_games):
        # REVERSE INDEX, scraped table data is ascending chronological & we want most recent
        score = team_data[ team_score_key ][-1-i]
        score = [int(x) for x in re.findall(r'\d+', score)]     # regex out the two numbers & convert to int
        score = max(score) if team_data[win_loss_key][-1 - i] == "W" else min(score)

        reverse_idx = n_previous_games - i - 1
        worksheet.cell(column=col_offset + reverse_idx, row=runs_row, value=score)

# =============================================sub function =========================================================

def WriteProfile(pitcher_data, worksheet, row_offset=0, col_offset=0):

    worksheet.cell(column=col_offset, row=row_offset, value=pitcher_data['team_id'])
    worksheet.cell(column=col_offset + 1, row=row_offset, value=pitcher_data[ pitcher_name_key ])

    # Write previous decisions
    for i in range(n_previous_games):
        try:
            reverse_idx = n_previous_games - i + 1
            worksheet.cell(column=col_offset + reverse_idx, row=row_offset, value=pitcher_data[ pitcher_win_loss_key ][i])
        except:
            reverse_idx = n_previous_games - i + 1
            worksheet.cell(column=col_offset + reverse_idx, row=row_offset, value="NO DATA")

# =============================================sub function =========================================================

def WriteSuitabilityStats(team_data, worksheet, orig_col_offset=0, is_home=None):

    at_bats_total = team_data['batting']['Totals']['AB']
    runs_total = team_data['batting']['Totals']['R']

    home_at_bats_total = team_data['batting_home']['Totals']['AB']
    home_runs_total = team_data['batting_home']['Totals']['R']

    away_at_bats_total = team_data['batting_away']['Totals']['AB']
    away_runs_total = team_data['batting_away']['Totals']['R']

    # - - - - - - - - - - - -
    batting_home_Ws = team_data['batting_splits']['Home']['W']
    batting_home_Ls = team_data['batting_splits']['Home']['L']

    batting_away_Ws = team_data['batting_splits']['Away']['W']
    batting_away_Ls = team_data['batting_splits']['Away']['L']

    # - - - - - - - - - - - -
    try:
        pitcher_home_Ws = team_data['pitcher_deep']['By Breakdown']['Home']['W']
    except:
        pitcher_home_Ws = "NO DATA"
    try:
        pitcher_home_Ls = team_data['pitcher_deep']['By Breakdown']['Home']['L']
    except:
        pitcher_home_Ls = "NO DATA"


    pitcher_away_Ws = team_data['pitcher_deep']['By Breakdown']['Away']['W']
    pitcher_away_Ls = team_data['pitcher_deep']['By Breakdown']['Away']['L']

    # ----------------------------------------------------------------------

    # - - - - - - - - - - - - FORM

    col_offset = orig_col_offset + (1 if is_home else 0)

    worksheet.cell(column=col_offset, row=batting_total_row, value=at_bats_total)
    worksheet.cell(column=col_offset + suitability_stats_col_offset, row=batting_total_row, value=runs_total)

    worksheet.cell(column=col_offset, row=batting_home_row, value=home_at_bats_total)
    worksheet.cell(column=col_offset + suitability_stats_col_offset, row=batting_home_row, value=home_runs_total)

    worksheet.cell(column=col_offset, row=batting_away_row, value=away_at_bats_total)
    worksheet.cell(column=col_offset + suitability_stats_col_offset, row=batting_away_row, value=away_runs_total)

    # - - - - - - - - - - - - SUITABILITY - Batting

    worksheet.cell(column=col_offset, row=team_home_row, value=batting_home_Ws)
    worksheet.cell(column=col_offset + suitability_stats_col_offset, row=team_home_row, value=batting_home_Ls)

    worksheet.cell(column=col_offset, row=team_away_row, value=batting_away_Ws)
    worksheet.cell(column=col_offset + suitability_stats_col_offset, row=team_away_row, value=batting_away_Ls)

    # - - - - - - - - - - - - SUITABILITY - Team

    worksheet.cell(column=col_offset, row=pitcher_home_row, value=pitcher_home_Ws)
    worksheet.cell(column=col_offset + suitability_stats_col_offset, row=pitcher_home_row, value=pitcher_home_Ls)

    worksheet.cell(column=col_offset, row=pitcher_away_row, value=pitcher_away_Ws)
    worksheet.cell(column=col_offset + suitability_stats_col_offset, row=pitcher_away_row, value=pitcher_away_Ls)


def WritePitcherMatchupStats(team_data, worksheet, col_offset=0, matchup_row_offset=0):

    matchup_row = matchup_stats_row + matchup_row_offset

    pitcher_total_era = team_data['pitcher_deep']['Overall']['Total']['ERA']
    pitcher_ip_against_current = team_data['pitcher_deep']['Current Opponent Data']['IP']
    pitcher_era_against_current = team_data['pitcher_deep']['Current Opponent Data']['ERA']

    # - - - - - - - - - - - - MATCHUP - Pitcher ERA

    worksheet.cell(column=col_offset + matchup_stats_col_start_offset, row=matchup_row, value=pitcher_total_era)

    # - - - - - - - - - - - - MATCHUP
    if pitcher_ip_against_current is not None and pitcher_era_against_current is not None:
        worksheet.cell(column=col_offset + matchup_stats_col_start_offset + matchup_stats_col_offset,
                       row=matchup_row,
                       value=pitcher_ip_against_current)
        worksheet.cell(column=col_offset + matchup_stats_col_start_offset + matchup_stats_col_offset * 2,
                       row=matchup_row,
                       value=pitcher_era_against_current)
    else:
        worksheet.cell(column=col_offset + matchup_stats_col_start_offset + matchup_stats_col_offset,
                       row=matchup_row,
                       value="NO DATA")
        worksheet.cell(column=col_offset + matchup_stats_col_start_offset + matchup_stats_col_offset * 2,
                       row=matchup_row,
                       value="NO DATA")

# ================================================================================================================


def WriteScrapedDataToExcel(data):

    base = openpyxl.load_workbook(filename="worksheets/RUN-SHEET-BASE.xlsx")

    # = = = = = = = = = = = = = = = = = = = = = = RATINGS  = = = = = = = = = = = = = = = = = =

    all_ratings = [base['RATINGS - 1'], base['RATINGS - 2']]

    print(base.sheetnames)

    total_col_offset = -1
    sheet_idx = -1

    for i, game in enumerate(data):

        if i % n_games_per_sheet == 0:                                      # Aneryin spreads games over sheets, so do this too...
            sheet_idx += 1                                                  # Go to the next sheet
            total_col_offset = ratings_data_column_start + 1                # Start at the beginning column

        if sheet_idx >= len(all_ratings):
            break

        ratings = all_ratings[sheet_idx]

        home = game['HOME']
        away = game['AWAY']

        try:
            WritePitcherMatchupStats(away, ratings, total_col_offset, matchup_row_offset=0)  # no vertical offset
        except:
            pass

        try:
            WritePitcherMatchupStats(home, ratings, total_col_offset, matchup_row_offset=1)  # no vertical offset
        except:
            pass

        # Away side
        try:
            WriteHeader(away, ratings, total_col_offset - 1, is_home=False)
        except:
            pass
        try:
            WriteResult(away, ratings, total_col_offset)
        except:
            pass
        try:
            WriteRuns(away, ratings, total_col_offset)
        except:
            pass
        try:
            WriteSuitabilityStats(away, ratings, total_col_offset, is_home=False)
        except:
            pass

        total_col_offset += n_previous_games + col_offset_home_away

        # home side
        try:
            WriteHeader(home, ratings, total_col_offset - 1, is_home=True)
        except:
            pass
        try:
            WriteResult(home, ratings, total_col_offset)
        except:
            pass
        try:
            WriteRuns(home, ratings, total_col_offset)
        except:
            pass
        try:
            WriteSuitabilityStats(home, ratings, total_col_offset, is_home=True)  # no vertical offset
        except:
            pass



        total_col_offset += n_previous_games + col_offset_games

    # = = = = = = = = = = = = = = = = = = = = = PROFILING  = = = = = = = = = = = = = = = = = = = =

    profiling = base['PROFILING']

    pitcher_row = -1
    pitcher_col = -1

    for i, game in enumerate(data):

        if i % n_games_per_sheet == 0:  # Aneryin spreads games over sheets, so do this too...
            pitcher_row = profile_data_row_start
            pitcher_col = profile_data_column_start + (profile_neat_layout_column_jump * round(i/n_games_per_sheet, 0))

        try:
            home = game['HOME']['pitcher']
            away = game['AWAY']['pitcher']
        except:
            pitcher_row += profile_row_offset_home_away
            pitcher_row += profile_row_offset_game
            continue

        WriteProfile(away, profiling, pitcher_row, pitcher_col)

        pitcher_row += profile_row_offset_home_away

        WriteProfile(home, profiling, pitcher_row, pitcher_col)

        pitcher_row += profile_row_offset_game

    # = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = = =

    base.save("excel/MLB_{}.xlsx".format(datetime.datetime.today().strftime('%Y-%m-%d')))


# ================================================================================================================
#
if __name__ == "__main__":

    with open('./todays_espn_game_data.json', 'r') as fp:
        data = json.load(fp)
        # pprint.pprint(data)
        WriteScrapedDataToExcel(data)